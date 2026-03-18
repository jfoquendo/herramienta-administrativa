import streamlit as st
import pandas as pd
import pikepdf
import re
import io
from openpyxl.styles import PatternFill

# --- CONFIGURACIÓN DE PÁGINA ---
st.set_page_config(
    page_title="Multi-Tool Administrativa 2026",
    page_icon="🔐",
    layout="wide"
)

# --- 1. BASE DE DATOS DE USUARIOS (Sencillo) ---
# Puedes agregar tantos como necesites siguiendo el formato "usuario": "contraseña"
USUARIOS_AUTORIZADOS = {
    "oquendo": "admin2026",
    "exssycortes": "Migajera2026**",
    "operador1": "clave123",
    "invitado": "654321"
}

def check_password():
    """Retorna True si el usuario introdujo credenciales válidas."""
    def login():
        user = st.session_state["username"]
        pw = st.session_state["password"]
        if user in USUARIOS_AUTORIZADOS and USUARIOS_AUTORIZADOS[user] == pw:
            st.session_state["password_correct"] = True
            del st.session_state["password"]
            del st.session_state["username"]
        else:
            st.session_state["password_correct"] = False

    if "password_correct" not in st.session_state:
        st.title("🔐 Acceso al Sistema")
        st.text_input("Usuario", key="username")
        st.text_input("Contraseña", type="password", key="password")
        st.button("Entrar", on_click=login)
        return False
    elif not st.session_state["password_correct"]:
        st.title("🔐 Acceso al Sistema")
        st.text_input("Usuario", key="username")
        st.text_input("Contraseña", type="password", key="password")
        st.button("Entrar", on_click=login)
        st.error("😕 Credenciales incorrectas")
        return False
    return True

# --- FUNCIONES DE LÓGICA ---
def limpiar_extremo(dato):
    if pd.isna(dato): return ""
    s = str(dato).strip()
    s = re.sub(r'\.0$', '', s)
    s = re.sub(r'\D', '', s)
    return s

def unlock_pdf(file_bytes, passwords):
    try:
        try:
            with pikepdf.open(io.BytesIO(file_bytes)) as pdf:
                out = io.BytesIO(); pdf.save(out)
                return True, "Liberado.", out.getvalue()
        except pikepdf.PasswordError:
            pass 
        for pw in passwords:
            try:
                with pikepdf.open(io.BytesIO(file_bytes), password=pw.strip()) as pdf:
                    out = io.BytesIO(); pdf.save(out)
                    return True, "Desbloqueado.", out.getvalue()
            except pikepdf.PasswordError: continue
        return False, "Contraseña incorrecta.", None
    except Exception as e: return False, f"Error: {str(e)}", None

# --- INICIO DE LA APLICACIÓN ---
if check_password():
    
    # Menú Lateral
    with st.sidebar:
        st.title("🛠️ Menú Principal")
        opcion = st.radio("Herramientas:", ["📱 Teléfonos", "🔓 PDFs", "👥 Cruce", "📊 Organizador"])
        st.markdown("---")
        if st.button("Cerrar Sesión"):
            st.session_state["password_correct"] = False
            st.rerun()

    # --- PESTAÑA 1: TELÉFONOS ---
    if opcion == "📱 Teléfonos":
        st.header("📱 Extractor WhatsApp")
        archivo = st.file_uploader("Subir Excel", type=["xlsx", "xls"])
        if archivo:
            df_temp = pd.read_excel(archivo, nrows=5)
            c1, c2 = st.columns(2)
            with c1: col_obs = st.selectbox("Columna Observaciones:", df_temp.columns)
            with c2: col_tel = st.selectbox("Columna Teléfonos:", df_temp.columns)

            if st.button("Procesar"):
                bar = st.progress(0)
                df = pd.read_excel(archivo)
                cats = {'Cédula':['cedula'],'Firma':['firma'],'Foto':['foto'],'Carta':['carta'],'Cesantias':['cesantias'],'EPS':['eps'],'ADRES':['adres'],'Bancario':['bancario','cuenta'],'Incompleto':['incompleto'],'Acta de Grado':['acta'],"Ruaf":['ruaf']}
                res = {c: [] for c in cats}
                for i, fila in df.iterrows():
                    obs = str(fila.get(col_obs, '')).lower()
                    tel = fila.get(col_tel)
                    if pd.isna(obs) or obs in ['nan','ok',''] or pd.isna(tel): continue
                    num = f"A,57{limpiar_extremo(tel)}"
                    for c, pws in cats.items():
                        if any(p in obs for p in pws): res[c].append(num)
                    bar.progress((i+1)/len(df))
                
                max_l = max(len(v) for v in res.values()) if res.values() else 0
                for c in res: res[c] += [None]*(max_l - len(res[c]))
                df_res = pd.DataFrame(res)
                
                st.subheader("👀 Vista Previa")
                st.dataframe(df_res, use_container_width=True)
                
                out = io.BytesIO()
                df_res.to_excel(out, index=False)
                st.download_button("📥 Descargar", out.getvalue(), "Telefonos.xlsx")

    # --- PESTAÑA 2: PDFs ---
    elif opcion == "🔓 PDFs":
        st.header("🔓 Desbloqueo Masivo PDF")
        pws = st.text_input("Contraseñas (separadas por coma)")
        p_files = st.file_uploader("PDFs", type="pdf", accept_multiple_files=True)
        if p_files and st.button("Ejecutar"):
            list_p = [p.strip() for p in pws.split(',')] if pws else []
            bar = st.progress(0)
            for i, pf in enumerate(p_files):
                ok, msg, content = unlock_pdf(pf.read(), list_p)
                if ok:
                    st.success(f"✅ {pf.name}")
                    st.download_button(f"Descargar {pf.name}", content, f"unlocked_{pf.name}", key=f"p_{i}")
                else: st.error(f"❌ {pf.name}: {msg}")
                bar.progress((i+1)/len(p_files))

    # --- PESTAÑA 3: CRUCE ---
    # --- PESTAÑA 3: CRUCE ---
    elif opcion == "👥 Cruce":
        st.header("👥 Cruce de Empleados")
        st.write("Cruce masivo con resaltado de duplicados y detección de faltantes.")
        
        c1, c2 = st.columns(2)
        with c1: m_f = st.file_uploader("1. Maestro (Activos)", type="xlsx", key="m_cruce")
        with c2: b_f = st.file_uploader("2. Lista Búsqueda (IDs)", type="xlsx", key="b_cruce")
        
        if m_f and b_f:
            # Lectura rápida de cabeceras para el selector
            df_m_h = pd.read_excel(m_f, skiprows=1, nrows=0)
            sel_id = st.selectbox("Selecciona columna ID en Maestro:", df_m_h.columns)
            
            # El botón de ejecución
            ejecutar = st.button("🚀 Iniciar Cruce de Datos")

            if ejecutar:
                # Usamos st.status para mostrar el progreso de forma profesional
                with st.status("Procesando cruce de información...", expanded=True) as status:
                    
                    st.write("📥 Cargando archivos Excel...")
                    df_a = pd.read_excel(m_f, skiprows=1)
                    cols_o = df_a.columns.tolist()
                    df_b = pd.read_excel(b_f, header=None)
                    
                    st.write("🧹 Limpiando y normalizando IDs...")
                    df_a['ID_L'] = df_a[sel_id].apply(limpiar_extremo)
                    # Convertimos la lista de búsqueda en un set para velocidad (O(1))
                    ceds = set([limpiar_extremo(v) for v in df_b.values.flatten() if limpiar_extremo(v) != ""])
                    
                    st.write("🔎 Buscando coincidencias...")
                    enc = df_a[df_a['ID_L'].isin(ceds)].copy()
                    enc = enc[cols_o] # Restauramos el orden de columnas original
                    
                    id_hallados = set(enc['ID_L'])
                    faltantes = [c for c in ceds if c not in id_hallados]
                    
                    st.write("🎨 Aplicando formato y colores...")
                    out_e = io.BytesIO()
                    with pd.ExcelWriter(out_e, engine='openpyxl') as writer:
                        enc.to_excel(writer, index=False)
                        ws = writer.book.active
                        fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        # Identificamos duplicados en la columna original
                        dup_mask = enc[sel_id].duplicated(keep=False).tolist()
                        for i, is_dup in enumerate(dup_mask):
                            if is_dup:
                                for cell in ws[i+2]: # i+2 por encabezado y base 1
                                    cell.fill = fill
                    
                    status.update(label="✅ Cruce completado con éxito!", state="complete", expanded=False)

                # --- RESULTADOS (Solo se muestran al terminar) ---
                st.markdown("---")
                col_m1, col_m2 = st.columns(2)
                with col_m1:
                    st.metric("Cédulas Encontradas", len(enc))
                with col_m2:
                    st.metric("Cédulas Faltantes", len(faltantes))

                st.subheader("👀 Vista Previa de Resultados (Primeros 100)")
                st.dataframe(enc.head(100), use_container_width=True)
                
                # Botones de descarga
                c_btn1, c_btn2 = st.columns(2)
                with c_btn1:
                    st.download_button(
                        label="📥 Descargar EXISTENTES (Excel)",
                        data=out_e.getvalue(),
                        file_name="Cruce_Existentes_Color.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                with c_btn2:
                    if faltantes:
                        df_falt = pd.DataFrame(faltantes, columns=['ID_No_Encontrado'])
                        out_f = io.BytesIO()
                        df_falt.to_excel(out_f, index=False)
                        st.download_button(
                            label="📥 Descargar FALTANTES (Excel)",
                            data=out_f.getvalue(),
                            file_name="Cruce_Faltantes.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                st.balloons()
    # --- PESTAÑA 4: ORGANIZADOR ---
    elif opcion == "📊 Organizador":
        st.header("📊 Organizador Excel")
        c1, c2 = st.columns(2)
        with c1: d_f = st.file_uploader("Datos", type="xlsx")
        with c2: o_f = st.file_uploader("Orden", type="xlsx")
        if d_f and o_f:
            id_d = st.selectbox("ID Datos:", pd.read_excel(d_f, nrows=0).columns)
            id_o = st.selectbox("ID Orden:", pd.read_excel(o_f, nrows=0).columns)
            if st.button("Organizar"):
                bar = st.progress(50)
                df_d = pd.read_excel(d_f)
                cols = df_d.columns.tolist()
                df_o = pd.read_excel(o_f)
                df_d[id_d] = df_d[id_d].astype(str).str.strip()
                df_o[id_o] = df_o[id_o].astype(str).str.strip()
                
                df_res = pd.merge(df_o[[id_o]], df_d, left_on=id_o, right_on=id_d, how='left')
                df_res = df_res[cols]
                bar.progress(100)
                
                st.subheader("📊 Resultado Organizado")
                st.dataframe(df_res, use_container_width=True)
                
                out = io.BytesIO()
                df_res.to_excel(out, index=False)
                st.download_button("📥 Descargar", out.getvalue(), "Organizado.xlsx")
